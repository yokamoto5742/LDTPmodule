import os
import configparser
from datetime import datetime
from openpyxl import load_workbook
from services.barcode_service import BarcodeService


class TreatmentPlanGenerator:
    """
    生活習慣病療養計画書を生成するサービスクラス
    """

    def __init__(self):
        """治療計画書生成サービスの初期化"""
        # 設定ファイルの読み込み
        self.config = configparser.ConfigParser()
        self.config.read('config.ini', encoding='utf-8')
        self.template_path = self.config.get("Paths", "template_path")
        self.output_path = self.config.get("Paths", "output_path")
        self.document_number = self.config.get('Document', 'document_number', fallback='39221')

        # バーコードサービスの初期化
        self.barcode_service = BarcodeService()

    def generate_plan(self, patient_info, file_name=None):
        """
        治療計画書を生成する

        Args:
            patient_info (PatientInfo): 患者情報オブジェクト
            file_name (str, optional): ファイル名（指定しない場合は自動生成）

        Returns:
            str: 生成されたファイルのパス
        """
        try:
            # ファイル名の生成
            current_time = datetime.now().strftime("%H%M%S")
            patient_id = str(patient_info.patient_id).zfill(9)
            department_id = str(patient_info.department_id).zfill(3)
            doctor_id = str(patient_info.doctor_id).zfill(5)
            issue_date = patient_info.issue_date.strftime("%Y%m%d")

            # バーコードデータの生成
            barcode_data = self.barcode_service.generate_barcode_data(
                patient_info.patient_id,
                patient_info.department_id,
                patient_info.doctor_id,
                patient_info.issue_date
            )

            if file_name is None:
                file_name = f"{patient_id}{self.document_number}{department_id}{doctor_id}{issue_date}{current_time}.xlsm"

            file_path = os.path.join(self.output_path, file_name)

            # Excelテンプレートの読み込み
            workbook = load_workbook(self.template_path, keep_vba=True)
            common_sheet = workbook["共通情報"]

            # 初回用と継続用のシートを取得
            initial_sheet = workbook["初回用"]
            continuous_sheet = workbook["継続用"]

            # 共通情報シートに患者情報を入力
            self._populate_common_sheet(common_sheet, patient_info)

            # 両方のシートにバーコードを追加
            self.barcode_service.add_barcode_to_sheet(initial_sheet, barcode_data)
            self.barcode_service.add_barcode_to_sheet(continuous_sheet, barcode_data)

            # ファイルの保存
            workbook.save(file_path)

            # 適切なシートをアクティブにする処理
            self._set_active_sheet(file_path, patient_info.creation_count)

            return file_path

        except Exception as e:
            raise Exception(f"治療計画書の生成中にエラーが発生しました: {str(e)}")

    def _populate_common_sheet(self, common_sheet, patient_info):
        """
        Excelの共通情報シートに患者情報を入力する

        Args:
            common_sheet: openpyxlのワークシートオブジェクト
            patient_info (PatientInfo): 患者情報オブジェクト
        """
        # 基本情報
        common_sheet["B2"] = patient_info.patient_id
        common_sheet["B3"] = patient_info.patient_name
        common_sheet["B4"] = patient_info.kana
        common_sheet["B5"] = patient_info.gender
        common_sheet["B6"] = patient_info.birthdate
        common_sheet["B7"] = patient_info.issue_date
        common_sheet["B8"] = patient_info.doctor_id
        common_sheet["B9"] = patient_info.doctor_name
        common_sheet["B10"] = patient_info.department_id
        common_sheet["B11"] = patient_info.department
        common_sheet["B12"] = patient_info.main_diagnosis
        common_sheet["B13"] = patient_info.creation_count
        common_sheet["B14"] = patient_info.target_weight
        common_sheet["B15"] = patient_info.sheet_name

        # 目標情報
        common_sheet["B16"] = patient_info.target_bp
        common_sheet["B17"] = patient_info.target_hba1c
        common_sheet["B18"] = patient_info.goal1
        common_sheet["B19"] = patient_info.goal2
        common_sheet["B20"] = patient_info.target_achievement

        # 食事指導情報
        common_sheet["B21"] = patient_info.diet1
        common_sheet["B22"] = patient_info.diet2
        common_sheet["B23"] = patient_info.diet3
        common_sheet["B24"] = patient_info.diet4
        common_sheet["B38"] = patient_info.diet_comment

        # 運動指導情報
        common_sheet["B25"] = patient_info.exercise_prescription
        common_sheet["B26"] = patient_info.exercise_time
        common_sheet["B27"] = patient_info.exercise_frequency
        common_sheet["B28"] = patient_info.exercise_intensity
        common_sheet["B29"] = patient_info.daily_activity
        common_sheet["B39"] = patient_info.exercise_comment

        # 生活習慣関連情報
        common_sheet["B30"] = patient_info.nonsmoker
        common_sheet["B31"] = patient_info.smoking_cessation
        common_sheet["B32"] = patient_info.other1
        common_sheet["B33"] = patient_info.other2

        # 受診勧奨情報
        common_sheet["B34"] = patient_info.ophthalmology
        common_sheet["B35"] = patient_info.dental
        common_sheet["B36"] = patient_info.cancer_screening

        # その他の情報
        common_sheet["B37"] = patient_info.issue_date_age

    def _set_active_sheet(self, file_path, creation_count):
        """
        作成回数に応じて適切なシートをアクティブにする

        Args:
            file_path (str): Excelファイルのパス
            creation_count (int): 作成回数
        """
        wb = load_workbook(file_path, read_only=False, keep_vba=True)
        ws_common = wb["共通情報"]
        ws_common.sheet_view.tabSelected = False

        # 全てのシートの選択状態をリセット
        for sheet in wb.worksheets:
            sheet.sheet_view.tabSelected = False

        # 作成回数に応じて適切なシートをアクティブにする
        if creation_count == 1:
            ws_plan = wb["初回用"]
        else:
            ws_plan = wb["継続用"]

        ws_plan.sheet_view.tabSelected = True
        wb.active = ws_plan

        wb.save(file_path)

    def open_file(self, file_path):
        """
        生成したExcelファイルを開く

        Args:
            file_path (str): 開くファイルのパス
        """
        os.startfile(file_path)


class DropdownItems:
    """
    ドロップダウンメニューの項目を管理するクラス
    """

    def __init__(self):
        """ドロップダウンアイテムマネージャーの初期化"""
        self.items = {
            'target_achievement': ['概ね達成', '概ね70%達成', '概ね50%達成', '未達成', '(空欄)'],
            'diet': ['食事量を適正にする', "塩分量を適正にする", '水分摂取量を増やす', '食物繊維の摂取量を増やす',
                     'ゆっくり食べる', '間食を減らす', 'アルコールを控える', '脂肪の多い食品や甘い物を控える',
                     '揚げ物や炒め物などを減らす', '1日3食を規則正しくとる', '今回は指導の必要なし', '(空欄)'],
            'exercise_prescription': ['ウォーキング', 'ストレッチ体操', '筋力トレーニング', '自転車', '畑仕事',
                                      '今回は指導の必要なし', '(空欄)'],
            'exercise_time': ['10分', '20分', '30分', '60分', '(空欄)'],
            'exercise_frequency': ['毎日', '週に5日', '週に3日', '週に2日', '(空欄)'],
            'exercise_intensity': ['息が弾む程度', 'ニコニコペース', '少し汗をかく程度', '息切れしない程度', '(空欄)'],
            'daily_activity': ['3000歩', '5000歩', '6000歩', '8000歩', '10000歩', 'ストレッチ運動を主に行う', '(空欄)'],
        }

    def get_options(self, key):
        """
        指定されたキーのオプションリストを取得する

        Args:
            key (str): オプションのカテゴリキー

        Returns:
            list: オプションのリスト
        """
        return self.items.get(key, [])

    def add_item(self, key, options):
        """
        新しいオプションカテゴリを追加する

        Args:
            key (str): 新しいカテゴリのキー
            options (list): オプションのリスト
        """
        self.items[key] = options

    def create_dropdown(self, key, label, width):
        """
        ドロップダウンフィールドを作成する

        Args:
            key (str): 選択肢のカテゴリキー
            label (str): ドロップダウンのラベル
            width (int): ドロップダウンの幅

        Returns:
            ft.Dropdown: 作成されたドロップダウンオブジェクト
        """
        import flet as ft
        options = [ft.dropdown.Option(str(item)) for item in self.get_options(key)]

        return ft.Dropdown(
            label=label,
            width=width,
            options=options,
            border_color=ft.colors.ON_SURFACE_VARIANT,
            focused_border_color=ft.colors.PRIMARY,
            text_style=ft.TextStyle(size=13),
            color=ft.colors.ON_SURFACE,
        )
